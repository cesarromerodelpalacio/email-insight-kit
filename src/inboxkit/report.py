"""Panel HTML (con SVG) + Excel de la analitica de bandeja."""
from __future__ import annotations

from pathlib import Path

from .analyze import InboxReport


def _bars(pairs, width=680, bar_h=26, gap=10, pad=200, label_max=26):
    top = pairs[:10]
    if not top:
        return ""
    mx = max((v for _, v in top), default=1) or 1
    h = len(top) * (bar_h + gap) + gap
    rows, y = [], gap
    for name, val in top:
        w = int((width - pad - 70) * (val / mx))
        lbl = str(name)[:label_max]
        rows.append(
            f'<text x="0" y="{y + bar_h*0.7:.0f}" fill="#8aa0b2" font-size="13">{lbl}</text>'
            f'<rect x="{pad}" y="{y}" width="{max(w,2)}" height="{bar_h}" rx="6" fill="#2dd4bf"/>'
            f'<text x="{pad + max(w,2) + 8}" y="{y + bar_h*0.7:.0f}" fill="#eaf2f7" font-size="12" font-weight="700">{val}</text>')
        y += bar_h + gap
    return f'<svg width="{width}" height="{h}" viewBox="0 0 {width} {h}" font-family="Segoe UI">{"".join(rows)}</svg>'


def _hour_bars(by_hour, width=680, h=140):
    mx = max((v for _, v in by_hour), default=1) or 1
    bw = width / 24
    bars = ""
    for hr, v in by_hour:
        bh = int((h - 24) * (v / mx))
        x = hr * bw
        bars += (f'<rect x="{x+3:.0f}" y="{h-20-bh}" width="{bw-6:.0f}" height="{bh}" rx="3" fill="#0ea5a0"/>'
                 f'<text x="{x+bw/2:.0f}" y="{h-6}" fill="#8aa0b2" font-size="9" text-anchor="middle">{hr}</text>')
    return f'<svg width="{width}" height="{h}" viewBox="0 0 {width} {h}" font-family="Segoe UI">{bars}</svg>'


def build_html(r: InboxReport) -> str:
    cards = (
        f'<div class="c"><div class="n">{r.total}</div><div class="l">Correos ({r.dias_analizados}d)</div></div>'
        f'<div class="c"><div class="n" style="color:#f59e0b">{r.unread}</div><div class="l">No leidos</div></div>'
        f'<div class="c"><div class="n">{r.unread_rate}%</div><div class="l">% sin leer</div></div>'
        f'<div class="c"><div class="n">{len(r.top_senders) and len(set(s for s,_ in r.top_senders))}</div><div class="l">Top remitentes</div></div>'
    )
    cats = "".join(f'<span class="chip">{c}: <b>{n}</b></span>' for c, n in r.categorias)
    prio = "".join(f"<tr><td class='k'>{s}</td><td>{subj[:70]}</td></tr>" for s, subj in r.prioridad) or "<tr><td colspan=2 style='color:#8aa0b2'>Sin no leidos importantes. Bandeja bajo control.</td></tr>"
    style = """
  body{font-family:'Segoe UI',Arial,sans-serif;background:#0a0f15;color:#eaf2f7;margin:0;padding:32px}
  .wrap{max-width:900px;margin:0 auto} h1{font-size:28px;margin:0}
  .sub{color:#8aa0b2;margin:6px 0 22px} h2{font-size:18px;margin:26px 0 10px}
  .cards{display:flex;gap:14px;flex-wrap:wrap} .c{flex:1;min-width:150px;background:#16212e;border:1px solid #26384a;border-radius:14px;padding:18px}
  .c .n{font-size:30px;font-weight:800;color:#2dd4bf} .c .l{color:#8aa0b2;font-size:12px;margin-top:4px;text-transform:uppercase;letter-spacing:.04em}
  .panel{background:#111a24;border:1px solid #26384a;border-radius:14px;padding:18px}
  .chip{display:inline-block;background:#16212e;border:1px solid #26384a;border-radius:999px;padding:6px 14px;margin:0 8px 8px 0;font-size:13px}
  table{width:100%;border-collapse:collapse;font-size:13px} td{padding:8px;border-bottom:1px solid #1f2e3d} td.k{color:#2dd4bf;font-weight:600;white-space:nowrap}
  .foot{color:#8aa0b2;font-size:12px;margin-top:24px;border-top:1px solid #26384a;padding-top:14px}
"""
    return (
        f'<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Email Insight</title><style>{style}</style></head><body><div class="wrap">'
        f'<h1>Panel de bandeja</h1><div class="sub">Generado {r.generated_at} · ultimos {r.dias_analizados} dias · 100% local</div>'
        f'<div class="cards">{cards}</div>'
        f'<h2>Quien te inunda (top remitentes)</h2><div class="panel">{_bars(r.top_senders)}</div>'
        f'<h2>Cuando llegan (por hora)</h2><div class="panel">{_hour_bars(r.by_hour)}</div>'
        f'<h2>Tipos de correo</h2><div>{cats}</div>'
        f'<h2>Prioridad: no leidos que requieren accion</h2><div class="panel"><table>{prio}</table></div>'
        f'<div class="foot">Email Insight Kit &middot; Cesar Romero del Palacio &middot; Todo procesado en local, nada sube a la nube.</div>'
        f'</div></body></html>'
    )


def write_html(r: InboxReport, out_path) -> Path:
    p = Path(out_path); p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(build_html(r), encoding="utf-8"); return p


def write_excel(r: InboxReport, out_path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    p = Path(out_path); p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    def _sheet(title, headers, rows):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] or title != "Resumen" else wb.active
        ws.title = title
        ws.append(headers)
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor="0F1720"); c.font = Font(color="FFFFFF", bold=True)
        for row in rows:
            ws.append(list(row))
        return ws
    _sheet("Resumen", ["Metrica", "Valor"], [("Total", r.total), ("No leidos", r.unread), ("% sin leer", r.unread_rate), ("Dias", r.dias_analizados)])
    _sheet("Top remitentes", ["Remitente", "Correos"], r.top_senders)
    _sheet("Dominios", ["Dominio", "Correos"], r.top_domains)
    _sheet("Categorias", ["Categoria", "Correos"], r.categorias)
    _sheet("Prioridad", ["Remitente", "Asunto"], r.prioridad)
    wb.save(p); return p
